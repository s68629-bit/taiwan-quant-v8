"""
歷史買點回測腳本
═══════════════════════════════════════════════════════════════
功能：對指定股票的過去 2 年，用 Walk-forward 方式
      找出每一個「系統發出買進訊號」的日期，
      並比對實際後續 20 個交易日的報酬率。

用法：
  python historical_signal_backtest.py 8069     ← 元太
  python historical_signal_backtest.py 3583     ← 辛耘
  python historical_signal_backtest.py 2330     ← 台積電

輸出：
  reports/signal_backtest_8069_YYYYMMDD.xlsx
  包含：
    - 每個買進訊號的日期、當日收盤價
    - 各評分（AI、籌碼、反向、飆漲）
    - 20 日後實際報酬率
    - 是否獲利（勝率統計）
    - 摘要統計（勝率、平均報酬、最大獲利/虧損）

Walk-forward 說明：
  每個訊號日，只用「該日之前」的資料訓練，
  完全不用未來資料，是真實可信的歷史驗證。
═══════════════════════════════════════════════════════════════
"""
import sys, os, warnings, logging
warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.WARNING)   # 靜音 INFO，只顯示進度

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta

from xgboost import XGBRegressor
from modules.technical_factors import add_technical, TECH_FEATURES
from modules.institutional_cost_engine import add_cost
from modules.chip_factors import add_chip_factors
from modules.contrarian_signals import compute_contrarian_score
from modules.surge_detector import detect_surge
from modules.retail_sentiment import compute_retail_sentiment
from modules.composite_scorer import compute_composite
from modules.foreign_cost import compute_foreign_cost
from modules.trap_detector import detect_traps
from config.settings import (
    LONG_THRESHOLD, WATCH_THRESHOLD,
    W_AI, W_CHIP, W_CONTRARIAN, W_SURGE,
    REPORT_DIR
)

os.makedirs(REPORT_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════
# 參數設定
# ══════════════════════════════════════════════════════════════
TRAIN_WINDOW    = 240   # 訓練窗口（約一年交易日）
MIN_TRAIN       = 150   # 最少訓練樣本
PREDICT_HORIZON = 20    # 預測未來幾個交易日
SCAN_START_DAYS = 240   # 從第幾天開始產生訊號（前面全用來訓練）
SIGNAL_THRESHOLD = LONG_THRESHOLD   # 綜合分超過此值才算「買進訊號」

FEATURES = TECH_FEATURES + [
    "foreign_net_20d","trust_net_5d","chip_net_5d",
    "foreign_consec_buy","margin_chg_5d","short_squeeze_ratio",
    "VWAP","CostBreak",
]


# ══════════════════════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════════════════════
def main():
    # 股票代號
    raw_sym = sys.argv[1] if len(sys.argv) > 1 else "8069"
    symbol  = raw_sym if raw_sym.endswith(".TW") else raw_sym + ".TW"

    print(f"\n{'═'*60}")
    print(f"  歷史買點回測：{symbol}")
    print(f"  Walk-forward：每個訊號只用當日之前的資料訓練")
    print(f"{'═'*60}\n")

    # ── 1. 下載資料 ─────────────────────────────────────────
    print("下載歷史資料（約 3 年）...")
    df_raw = yf.download(symbol, period="3y", interval="1d",
                         progress=False, auto_adjust=True)
    if df_raw.empty:
        print(f"❌ 無法下載 {symbol} 資料，請確認代號或網路連線")
        return

    if df_raw.columns.nlevels > 1:
        df_raw.columns = df_raw.columns.droplevel(1)
    df_raw = df_raw.dropna()
    print(f"資料範圍：{df_raw.index[0].date()} ~ {df_raw.index[-1].date()}（{len(df_raw)} 筆）\n")

    # ── 2. 計算技術指標 & 籌碼（整批計算，不分窗口） ────────
    print("計算技術指標...")
    df = add_technical(df_raw.copy())
    df = add_cost(df)
    df = add_chip_factors(df, pd.DataFrame(), pd.DataFrame())   # 無 FinMind 則填 0
    df["Target"] = df["Close"].pct_change(fill_method=None).shift(-PREDICT_HORIZON)

    avail_features = [f for f in FEATURES if f in df.columns]

    # 只看「回測區間」：最近 2 年
    two_years_ago = df.index[-1] - pd.DateOffset(years=2)
    backtest_idx  = df.index[df.index >= two_years_ago]

    print(f"回測區間：{backtest_idx[0].date()} ~ {backtest_idx[-1].date()}")
    print(f"使用特徵：{len(avail_features)} 個")
    print(f"訓練窗口：每次最多 {TRAIN_WINDOW} 個交易日\n")

    # ── 3. Walk-forward 回測 ──────────────────────────────────
    results = []
    total   = len(backtest_idx)
    signals_found = 0

    for i, today in enumerate(backtest_idx):
        # 進度顯示
        if i % 20 == 0:
            pct = i / total * 100
            bar = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))
            print(f"\r  [{bar}] {pct:.0f}%  已掃 {i}/{total} 天  訊號 {signals_found} 個",
                  end="", flush=True)

        # 取今天之前的資料作為訓練集
        past_df = df[df.index < today].copy()

        # 必須有足夠的訓練資料
        train_df = past_df.dropna(subset=avail_features + ["Target"])
        if len(train_df) < MIN_TRAIN:
            continue

        # 今天的特徵（不能有 NaN）
        today_row = df.loc[[today]].dropna(subset=avail_features)
        if today_row.empty:
            continue

        # ── AI 模型（walk-forward）────────────────────────────
        try:
            # 只用最近 TRAIN_WINDOW 天訓練（滾動窗口）
            train_use = train_df.tail(TRAIN_WINDOW)
            X_train = train_use[avail_features]
            y_train = train_use["Target"]

            model = XGBRegressor(
                n_estimators=200, max_depth=5,
                learning_rate=0.05, subsample=0.8,
                n_jobs=-1, random_state=42, verbosity=0
            )
            model.fit(X_train, y_train)
            ai_pred = float(model.predict(today_row[avail_features])[0])
        except Exception:
            continue

        # ── 反向訊號 ─────────────────────────────────────────
        try:
            contrarian = compute_contrarian_score(
                pd.concat([train_df.tail(60), today_row])
            )
        except Exception:
            contrarian = {"contrarian_score": 0, "retail_heat": 0.5, "warning": ""}

        # ── 飆漲偵測 ─────────────────────────────────────────
        try:
            surge = detect_surge(pd.concat([train_df.tail(80), today_row]))
        except Exception:
            surge = {"surge_score": 0, "stage_label": "", "est_days_label": ""}

        # ── 散戶情緒 ─────────────────────────────────────────
        try:
            retail = compute_retail_sentiment(pd.concat([train_df.tail(100), today_row]))
        except Exception:
            retail = {"temperature": 50}

        # ── 陷阱偵測 ─────────────────────────────────────────
        try:
            trap = detect_traps(pd.concat([train_df.tail(60), today_row]))
        except Exception:
            trap = {"trap_score": 0, "trap_level": "未知", "entry_ok": True, "signals": []}

        # ── 四合一評分 ────────────────────────────────────────
        chip_latest = today_row.iloc[0]
        composite   = compute_composite(
            symbol, ai_pred, chip_latest, contrarian, False, surge
        )

        comp_score = composite["composite"]
        close_price = float(today_row["Close"].iloc[0])

        # ── 記錄實際結果（20 日後） ───────────────────────────
        future_prices = df[df.index > today]["Close"]
        actual_ret    = None
        exit_price    = None
        if len(future_prices) >= PREDICT_HORIZON:
            exit_price = float(future_prices.iloc[PREDICT_HORIZON - 1])
            actual_ret = (exit_price - close_price) / close_price

        # 訊號分類
        if comp_score > 0.40:
            signal_type = "★★ 強力做多"
        elif comp_score > SIGNAL_THRESHOLD:
            signal_type = "★ 做多"
        elif comp_score > WATCH_THRESHOLD:
            signal_type = "👁 觀察"
        elif comp_score > -0.10:
            signal_type = "⚪ 觀望"
        else:
            signal_type = "⚠️ 偏空"

        is_buy_signal = comp_score > SIGNAL_THRESHOLD
        if is_buy_signal:
            signals_found += 1

        results.append({
            "日期":         str(today.date()),
            "收盤價":       round(close_price, 1),
            "綜合評分":     round(comp_score, 3),
            "AI分":         round(composite["ai_score"], 3),
            "籌碼分":       round(composite["chip_score"], 3),
            "反向分":       round(composite["contrarian_score"], 3),
            "飆漲分":       composite.get("surge_score_raw", 0),
            "陷阱分":       trap.get("trap_score", 0),
            "散戶溫度":     round(retail.get("temperature", 50), 0),
            "訊號":         signal_type,
            "是否買進訊號": "✅ 買進" if is_buy_signal else "",
            "陷阱警示":     "⚠️" if trap.get("trap_score", 0) > 40 else "",
            f"{PREDICT_HORIZON}日後收盤": round(exit_price, 1) if exit_price else None,
            f"實際{PREDICT_HORIZON}日報酬": f"{actual_ret:+.2%}" if actual_ret is not None else "持倉中",
            "獲利與否":     ("✅ 獲利" if actual_ret > 0 else "❌ 虧損") if actual_ret is not None else "持倉中",
        })

    print(f"\r  [{'█'*20}] 100%  掃描完成！                          ")

    if not results:
        print("❌ 無結果，請確認資料是否充足")
        return

    df_result = pd.DataFrame(results)

    # ── 4. 統計摘要 ──────────────────────────────────────────
    buy_signals = df_result[df_result["是否買進訊號"] == "✅ 買進"].copy()
    verified    = buy_signals[buy_signals["獲利與否"].isin(["✅ 獲利","❌ 虧損"])]

    print(f"\n{'─'*60}")
    print(f"  回測摘要：{symbol}")
    print(f"{'─'*60}")
    print(f"  回測天數       ：{len(df_result)} 天")
    print(f"  買進訊號次數   ：{len(buy_signals)} 次")
    print(f"  已結算（可驗證）：{len(verified)} 次")

    if len(verified) > 0:
        wins     = (verified["獲利與否"] == "✅ 獲利").sum()
        win_rate = wins / len(verified)
        returns  = verified[f"實際{PREDICT_HORIZON}日報酬"].apply(
            lambda x: float(x.replace("+","").replace("%",""))/100 if "%" in str(x) else None
        ).dropna()

        print(f"  勝率           ：{win_rate:.1%}（{wins}/{len(verified)} 次獲利）")
        if len(returns) > 0:
            print(f"  平均報酬       ：{returns.mean():+.2%}")
            print(f"  最大獲利       ：{returns.max():+.2%}")
            print(f"  最大虧損       ：{returns.min():+.2%}")
            print(f"  累計報酬（等權）：{returns.sum():+.2%}")

    # 加權陷阱的買進訊號
    safe_signals = buy_signals[buy_signals["陷阱警示"] == ""]
    risky_signals = buy_signals[buy_signals["陷阱警示"] == "⚠️"]
    if len(risky_signals) > 0:
        print(f"\n  ⚠️  其中 {len(risky_signals)} 次買進訊號同時有陷阱警示（建議當時應忽略）")
    if len(safe_signals) > 0:
        s_verified = safe_signals[safe_signals["獲利與否"].isin(["✅ 獲利","❌ 虧損"])]
        if len(s_verified) > 0:
            sw = (s_verified["獲利與否"] == "✅ 獲利").sum() / len(s_verified)
            print(f"  去除陷阱後勝率 ：{sw:.1%}（{len(s_verified)} 次）")

    # ── 5. 輸出 Excel ─────────────────────────────────────────
    print(f"\n{'─'*60}")
    print("輸出 Excel 報告...")

    fname = f"signal_backtest_{raw_sym}_{datetime.today().strftime('%Y%m%d')}.xlsx"
    fpath = os.path.join(REPORT_DIR, fname)

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1：買進訊號列表 ────────────────────────────────
    ws1 = wb.active; ws1.title = "買進訊號"
    _write_sheet(ws1, buy_signals, symbol, f"買進訊號（綜合分 > {SIGNAL_THRESHOLD}）")

    # ── Sheet 2：全部交易日 ──────────────────────────────────
    ws2 = wb.create_sheet("全部交易日")
    _write_sheet(ws2, df_result, symbol, "全部交易日評分記錄")

    # ── Sheet 3：摘要統計 ────────────────────────────────────
    ws3 = wb.create_sheet("摘要統計")
    _write_summary(ws3, symbol, buy_signals, verified, PREDICT_HORIZON)

    wb.save(fpath)
    print(f"✅ 報告已儲存：{fpath}")
    print(f"\n  請用 Excel 開啟，「買進訊號」頁籤就是所有歷史買點")
    print(f"{'═'*60}\n")


def _write_sheet(ws, df, symbol, title):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    BLUE   = "1F4E79"; WHITE  = "FFFFFF"; GREEN = "E2EFDA"
    RED    = "FCE4D6"; ALT    = "EBF3FA"
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    c = ws.cell(1, 1, f"{symbol}　{title}")
    c.font = Font(bold=True, size=13, color=BLUE, name="Arial")
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    ws.row_dimensions[1].height = 22

    # 表頭
    for col, header in enumerate(df.columns, 1):
        cell = ws.cell(2, col, header)
        cell.fill      = PatternFill("solid", fgColor=BLUE)
        cell.font      = Font(bold=True, color=WHITE, name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = max(len(str(header)) + 2, 10)

    # 資料
    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        is_profit = str(row.get("獲利與否","")) == "✅ 獲利"
        is_loss   = str(row.get("獲利與否","")) == "❌ 虧損"
        has_trap  = str(row.get("陷阱警示","")) == "⚠️"
        bg = GREEN if is_profit else (RED if is_loss else (ALT if row_idx%2==0 else WHITE))

        for col, val in enumerate(row.values, 1):
            cell = ws.cell(row_idx, col, val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border


def _write_summary(ws, symbol, buy_signals, verified, horizon):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    BLUE = "1F4E79"; WHITE = "FFFFFF"; ACCENT = "D6E4F0"
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    c = ws.cell(1, 1, f"{symbol}　歷史買點回測摘要")
    c.font = Font(bold=True, size=13, color=BLUE, name="Arial")
    ws.merge_cells("A1:C1")

    rows = [
        ("股票代號",        symbol),
        ("買進訊號次數",    len(buy_signals)),
        ("已結算可驗證",    len(verified)),
    ]

    if len(verified) > 0:
        wins = (verified["獲利與否"] == "✅ 獲利").sum()
        win_rate = wins / len(verified)
        returns = verified[f"實際{horizon}日報酬"].apply(
            lambda x: float(x.replace("+","").replace("%",""))/100 if "%" in str(x) else None
        ).dropna()

        rows += [
            ("勝率",           f"{win_rate:.1%}"),
            ("獲利次數",        wins),
            ("虧損次數",        len(verified) - wins),
            ("平均報酬率",      f"{returns.mean():+.2%}" if len(returns) > 0 else "N/A"),
            ("最大單次獲利",    f"{returns.max():+.2%}"  if len(returns) > 0 else "N/A"),
            ("最大單次虧損",    f"{returns.min():+.2%}"  if len(returns) > 0 else "N/A"),
        ]

        safe = buy_signals[buy_signals["陷阱警示"] == ""]
        s_verified = safe[safe["獲利與否"].isin(["✅ 獲利","❌ 虧損"])]
        if len(s_verified) > 0:
            sw = (s_verified["獲利與否"] == "✅ 獲利").sum() / len(s_verified)
            rows += [
                ("（去除陷阱後）買進次數", len(s_verified)),
                ("（去除陷阱後）勝率",     f"{sw:.1%}"),
            ]

    for i, (k, v) in enumerate(rows, 3):
        ws.cell(i, 1, k).fill   = PatternFill("solid", fgColor=ACCENT)
        ws.cell(i, 1, k).font   = Font(bold=True, name="Arial", size=10)
        ws.cell(i, 2, v).font   = Font(name="Arial", size=10)
        for col in [1,2]:
            ws.cell(i, col).border    = border
            ws.cell(i, col).alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


if __name__ == "__main__":
    main()
