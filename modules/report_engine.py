"""
Excel 回測報告產生器 v5
工作表：
  1. 選股排名（含三合一分數 + 散戶警示）
  2. 回測績效（淨值曲線圖）
  3. 交易明細（含停損標記）
  4. 特徵重要性（長條圖）
  5. 籌碼資料品質診斷
"""
import os, logging, datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from config.settings import REPORT_DIR

logger = logging.getLogger(__name__)
os.makedirs(REPORT_DIR, exist_ok=True)

# ── 色彩常數 ─────────────────────────────────────────────
C_BLUE      = "1F4E79"
C_LTBLUE    = "2E75B6"
C_ACCENT    = "D6E4F0"
C_ALT       = "EBF3FA"
C_GREEN_BG  = "E2EFDA"
C_RED_BG    = "FCE4D6"
C_YELLOW_BG = "FFF2CC"
C_WHITE     = "FFFFFF"
C_GRAY      = "F2F2F2"

thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def _fill(color): return PatternFill("solid", fgColor=color)
def _font(color=C_WHITE, bold=True, size=10):
    return Font(bold=bold, color=color, name="Arial", size=size)
def _cell(ws, r, c, val, bg=None, bold=False, align="center", size=10):
    cell = ws.cell(row=r, column=c, value=val)
    if bg: cell.fill = _fill(bg)
    cell.font      = Font(name="Arial", size=size, bold=bold,
                          color=C_WHITE if bg in (C_BLUE, C_LTBLUE) else "000000")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = BORDER
    return cell

def _header(ws, row, cols, widths, bg=C_BLUE):
    for i, (h, w) in enumerate(zip(cols, widths), 1):
        _cell(ws, row, i, h, bg=bg, bold=True)
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────
# 公開介面
# ─────────────────────────────────────────────────────────

def generate_report(scan_results, backtest_curve, backtest_trades,
                    backtest_metrics, val_results, chip_status=None):
    wb = Workbook()

    ws1 = wb.active;  ws1.title = "選股排名"
    _sheet_ranking(ws1, scan_results, val_results)

    ws2 = wb.create_sheet("回測績效")
    _sheet_backtest(ws2, backtest_metrics, backtest_curve)

    ws3 = wb.create_sheet("交易明細")
    _sheet_trades(ws3, backtest_trades)

    ws4 = wb.create_sheet("特徵重要性")
    _sheet_importance(ws4, val_results)

    ws5 = wb.create_sheet("籌碼品質診斷")
    _sheet_chip_quality(ws5, chip_status or {}, val_results)

    fname = f"quant_v5_report_{datetime.date.today().strftime('%Y%m%d_%H%M')}.xlsx"
    path  = os.path.join(REPORT_DIR, fname)
    wb.save(path)
    logger.info("報告輸出：%s", path)
    return path


# ── Sheet 1：選股排名 ─────────────────────────────────────

def _sheet_ranking(ws, results, val_results):
    c = ws.cell(1, 1, "Taiwan Quant Fund v5 ─ 三合一 AI 選股排名")
    c.font = Font(bold=True, size=14, color=C_BLUE, name="Arial")
    ws.merge_cells("A1:K1"); ws.row_dimensions[1].height = 24

    hdrs = ["排名","股票","綜合評分","AI分","籌碼分","反向分",
            "散戶過熱","訊號","IC值","OOS勝率","散戶警示"]
    wids = [6,12,10,8,8,8,10,10,8,10,30]
    _header(ws, 2, hdrs, wids)

    for rank, (sym, cdict) in enumerate(results, 1):
        vm  = (val_results.get(sym) or {}).get("metrics", {})
        row = rank + 2
        alt = rank % 2 == 0
        bg  = C_ALT if alt else C_WHITE
        heat = cdict.get("retail_heat", 0)
        heat_str = "▓"*int(heat*5) + "░"*(5-int(heat*5)) + f" {heat:.0%}"
        sig  = cdict.get("signal_icon","") + " " + cdict.get("signal","")
        data = [
            rank, sym,
            f"{cdict.get('composite',0):+.3f}",
            f"{cdict.get('ai_score',0):+.3f}",
            f"{cdict.get('chip_score',0):+.3f}",
            f"{cdict.get('contrarian_score',0):+.3f}",
            heat_str, sig,
            f"{vm.get('IC',0):.3f}",
            f"{vm.get('win_rate',0):.1%}",
            cdict.get("warning",""),
        ]
        for col, val in enumerate(data, 1):
            c = _cell(ws, row, col, val, bg=bg, align="left" if col > 8 else "center")
            # 散戶過熱高 → 紅底警示
            if col == 7 and heat > 0.60:
                c.fill = _fill(C_RED_BG)


# ── Sheet 2：回測績效 ─────────────────────────────────────

def _sheet_backtest(ws, metrics, curve_df):
    c = ws.cell(1,1,"回測績效摘要"); c.font = Font(bold=True,size=14,color=C_BLUE,name="Arial")
    ws.merge_cells("A1:C1")

    items = [
        ("總報酬率",      f"{metrics.get('total_return',0):.2%}"),
        ("年化報酬率",    f"{metrics.get('ann_return',0):.2%}"),
        ("夏普比率",      f"{metrics.get('sharpe',0):.2f}"),
        ("最大回撤",      f"{metrics.get('max_drawdown',0):.2%}"),
        ("回測月數",      str(metrics.get('n_months',0))),
        ("交易筆數",      str(metrics.get('n_trades',0))),
        ("月勝率",        f"{metrics.get('win_rate',0):.2%}"),
        ("平均每筆報酬",  f"{metrics.get('avg_trade_r',0):+.2%}"),
        ("觸停損次數",    str(metrics.get('stop_loss_hits',0))),
    ]
    for i,(k,v) in enumerate(items, 3):
        _cell(ws, i, 1, k, bg=C_ACCENT, bold=True, align="left")
        _cell(ws, i, 2, v, align="center")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14

    if not curve_df.empty:
        ws.cell(2,5,"月份"); ws.cell(2,6,"淨值")
        for i, row in enumerate(curve_df.itertuples(), 3):
            ws.cell(i,5,str(row.month))
            ws.cell(i,6,round(float(row.equity),4))
        chart = LineChart()
        chart.title = "投資組合淨值曲線"; chart.style = 10
        chart.width = 22; chart.height = 14
        data = Reference(ws, min_col=6, min_row=2, max_row=len(curve_df)+2)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, "D14")
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 12


# ── Sheet 3：交易明細 ─────────────────────────────────────

def _sheet_trades(ws, trades_df):
    c = ws.cell(1,1,"逐筆交易明細"); c.font = Font(bold=True,size=14,color=C_BLUE,name="Arial")
    ws.merge_cells("A1:F1")
    if trades_df.empty:
        ws.cell(2,1,"（無交易紀錄）"); return

    _header(ws, 2,
            ["月份","股票","綜合評分","實際報酬","停損觸發","獲利與否"],
            [12,12,12,14,12,12])

    for i, row in enumerate(trades_df.itertuples(), 3):
        hit = bool(row.hit)
        sl  = bool(getattr(row,"stop_loss_hit",False))
        bg  = C_RED_BG if sl else (C_GREEN_BG if hit else C_RED_BG)
        _cell(ws,i,1,str(row.month),        bg=bg)
        _cell(ws,i,2,row.symbol,             bg=bg)
        _cell(ws,i,3,f"{row.composite:+.3f}",bg=bg)
        _cell(ws,i,4,f"{row.actual_return:+.2%}",bg=bg)
        _cell(ws,i,5,"⛔ 是" if sl else "—",  bg=bg)
        _cell(ws,i,6,"✓ 獲利" if hit else "✗ 虧損", bg=bg)


# ── Sheet 4：特徵重要性 ───────────────────────────────────

def _sheet_importance(ws, val_results):
    c = ws.cell(1,1,"模型特徵重要性（各股平均）")
    c.font = Font(bold=True,size=14,color=C_BLUE,name="Arial")
    ws.merge_cells("A1:C1")

    all_imp = []
    for sym, out in val_results.items():
        imp = (out or {}).get("importance")
        if imp is not None and not imp.empty:
            all_imp.append(imp.set_index("feature")["importance"])
    if not all_imp: return

    avg = pd.concat(all_imp,axis=1).mean(axis=1).sort_values(ascending=False)
    _header(ws, 2, ["特徵名稱","平均重要性","面向"], [22,16,10])

    GROUP = {
        **{f:"籌碼" for f in ["foreign_net_20d","trust_net_5d","chip_net_5d",
                               "foreign_consec_buy","margin_chg_5d","short_squeeze_ratio"]},
        **{f:"技術" for f in ["MA20","MA60","MA_ratio","Momentum","Volatility",
                               "VolumeMA","VolumeRatio","RSI14","Dist_52W_High","PV_diverge"]},
        **{f:"成本" for f in ["VWAP","CostBreak"]},
    }
    GC = {"籌碼":C_ACCENT,"技術":C_ALT,"成本":C_YELLOW_BG}

    for i,(feat,score) in enumerate(avg.items(),3):
        g  = GROUP.get(feat,"其他")
        bg = GC.get(g, C_WHITE)
        _cell(ws,i,1,feat,        bg=bg, align="left")
        _cell(ws,i,2,f"{score:.4f}", bg=bg)
        _cell(ws,i,3,g,           bg=bg)

    chart = BarChart(); chart.type="bar"
    chart.title="平均特徵重要性"; chart.style=10
    chart.width=22; chart.height=18
    n = len(avg)
    data = Reference(ws, min_col=2, min_row=2, max_row=n+2)
    cats = Reference(ws, min_col=1, min_row=3, max_row=n+2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E2")


# ── Sheet 5：籌碼品質診斷 ─────────────────────────────────

def _sheet_chip_quality(ws, chip_status, val_results):
    c = ws.cell(1,1,"籌碼資料品質診斷")
    c.font = Font(bold=True,size=14,color=C_BLUE,name="Arial")
    ws.merge_cells("A1:D1")

    _header(ws, 2,
            ["股票","三大法人","融資融券","整體狀態"],
            [14,16,14,28])

    for i,(sym, real) in enumerate(chip_status.items(), 3):
        inst_ok   = real.get("chip_ok",   False)
        margin_ok = real.get("margin_ok", False)
        if inst_ok and margin_ok:
            status, bg = "✅ 資料完整，籌碼分析有效", C_GREEN_BG
        elif inst_ok or margin_ok:
            status, bg = "⚠️ 部分資料缺失，籌碼分析僅供參考", C_YELLOW_BG
        else:
            status, bg = "❌ 無籌碼資料，預測僅靠技術面", C_RED_BG

        _cell(ws,i,1,sym,                           bg=bg)
        _cell(ws,i,2,"✅ 正常" if inst_ok   else "❌ 缺失", bg=bg)
        _cell(ws,i,3,"✅ 正常" if margin_ok else "❌ 缺失", bg=bg)
        _cell(ws,i,4,status,                        bg=bg, align="left")

    ws.column_dimensions["D"].width = 32
